import asyncio
import json

from openpyxl import load_workbook
from rest_framework.response import Response
from rest_framework.views import APIView

from wildberries.models import Product
from wildberries.pydantic import CardPydantic
from wildberries.utils import make_request


class CardView(APIView):
    @staticmethod
    def get_card_info(value):
        page = asyncio.run(make_request(value))
        return CardView.get_objects(page, value)

    @staticmethod
    def get_cards_info(file):
        values = []
        wb = load_workbook(file)
        for sheet in wb.sheetnames:
            for row in wb[sheet].iter_rows(values_only=True):
                values.append(row[0])
        cards_info = [CardView.get_card_info(i) for i in values]
        return cards_info

    @staticmethod
    def get_objects(page, value):
        card = None
        try:
            products = json.dumps(page['data']['products'][0])
            card = CardPydantic.parse_raw(products)
            Product.objects.create(**card.dict())
        except IndexError:
            print(f'id {value} отсутствует на сайте wildberries.ru')
        if card:
            return card.dict()
        else:
            return {'error': f'id {value} отсутствует на сайте wildberries.ru'}

    def post(self, request, *args, **kwargs):
        data = None
        if 'file' in request.data and 'value' in request.data:
            return Response({'error': 'Одновременно отправлять поля '
                                      'file и value запрещено!'})
        elif 'file' in request.data:
            file = request.data['file']
            data = CardView.get_cards_info(file)
        elif 'value' in request.data:
            value = request.data['value']
            data = CardView.get_card_info(value)
        return Response(data)
