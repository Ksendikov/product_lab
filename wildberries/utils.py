import asyncio
import json

import aiohttp
from openpyxl import load_workbook

from wildberries.models import Card
from wildberries.pydantic import CardPydantic

URL = 'https://card.wb.ru/cards/detail'


async def make_request(value):
    async with aiohttp.ClientSession() as sessions:
        async with sessions.get(f'{URL}?nm={value}') as response:
            return await response.json(content_type=None)


def get_card_info(value):
    page = asyncio.run(make_request(value))
    return get_objects(page)


async def get_cards_info(file):
    queue = asyncio.Queue()
    task_list = []
    wb = load_workbook(file)
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows(values_only=True):
            task = asyncio.create_task(make_request(row[0]))
            task_list.append(task)
    await queue.join()
    values = await asyncio.gather(*task_list, return_exceptions=True)
    cards_info = [get_objects(value) for value in values]
    return cards_info


def get_objects(page):
    card = None
    try:
        products = json.dumps(page['data']['products'][0])
        card = CardPydantic.parse_raw(products)
    except IndexError:
        print('article  not found')
    if card:
        return card
    else:
        return {'error': 'article  not found'}


def update_or_create_card(card):
    defaults = {'brand': card.brand, 'title': card.title}
    response = Card.objects.update_or_create(
        article=card.article,
        defaults=defaults,
    )
    return response
